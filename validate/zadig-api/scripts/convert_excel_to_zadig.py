#!/usr/bin/env python3
"""
Convert OB application list Excel to Zadig standard format.

Usage:
    python scripts/convert_excel_to_zadig.py <input_excel> <output_excel>

Example:
    python scripts/convert_excel_to_zadig.py \
        "/Users/fngil35/Downloads/Onebackend 资源清单汇总.xlsx" \
        "/Users/fngil35/Downloads/zadig_projects_standard.xlsx"
"""

import sys
import re
import json
import argparse
from pathlib import Path
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def normalize_project_key(project_name: str) -> str:
    """
    Convert project_name to project_key.

    Rules:
    1. Convert to lowercase
    2. Replace spaces and special chars with hyphens
    3. Remove consecutive hyphens
    4. Strip leading/trailing hyphens

    Examples:
        "Management Portal" -> "management-portal"
        "Data Platform / Realtime Computing" -> "data-platform-realtime-computing"
    """
    if not project_name:
        return ""

    # Remove newlines and extra spaces
    cleaned = re.sub(r'\s+', ' ', str(project_name).strip())

    # Convert to lowercase
    cleaned = cleaned.lower()

    # Replace special chars with hyphens
    cleaned = re.sub(r'[^a-z0-9]+', '-', cleaned)

    # Remove consecutive hyphens
    cleaned = re.sub(r'-+', '-', cleaned)

    # Strip leading/trailing hyphens
    cleaned = cleaned.strip('-')

    return cleaned


def clean_project_name(project_name: str) -> str:
    """Clean project name by removing newlines and extra spaces."""
    if not project_name:
        return ""
    return re.sub(r'\s+', ' ', str(project_name).strip())


def clean_service_name(service: str) -> str:
    """Clean service name by removing extra spaces."""
    if not service:
        return ""
    return str(service).strip()


def parse_variable_yaml(variables_str: str) -> dict:
    """
    Parse variable string to dict.

    Supports:
    - JSON string: '{"key": "value"}'
    - Key-value pairs: "key=value, key2=value2"

    Returns empty dict if parsing fails.
    """
    if not variables_str:
        return {}

    variables_str = str(variables_str).strip()

    # Try JSON first
    try:
        return json.loads(variables_str)
    except json.JSONDecodeError:
        pass

    # Try key=value format
    result = {}
    for item in variables_str.split(','):
        item = item.strip()
        if '=' in item:
            key, value = item.split('=', 1)
            result[key.strip()] = value.strip()

    return result


def get_default_variables(service_row: dict, extra_vars: dict) -> dict:
    """
    Build default variable_yaml from service row data.

    Args:
        service_row: Dictionary containing service data
        extra_vars: Dictionary of extra variables from Excel

    Returns:
        Dictionary suitable for variable_yaml
    """
    variables = {}

    # Add namespace if available
    namespace = service_row.get('namespace', '')
    if namespace:
        variables['namespace'] = namespace

    # Add extra variables (only non-empty values)
    variable_keys = [
        'arch', 'component', 'computeType', 'cpuLimit', 'cpuRequest',
        'envType', 'imagePullPolicy', 'language', 'memoryLimit', 'memoryRequest',
        'metricsEnabled', 'metricsPath', 'nsGatewayHost', 'port', 'probePath',
        'registry', 'tag', 'team'
    ]

    for key in variable_keys:
        value = extra_vars.get(key)
        if value is not None and str(value).strip():
            variables[key] = str(value).strip()

    return variables


def read_source_excel(file_path: str) -> tuple[list, list]:
    """
    Read source Excel file and extract projects and services.

    Returns:
        Tuple of (projects_list, services_list)
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['OB 应用列表']

    # Read header
    header = [cell.value if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Map column names to indices
    col_map = {name: i for i, name in enumerate(header)}

    # Read data rows
    projects_dict = {}  # project_key -> project info
    services_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        project_name_raw = row[col_map.get('project_name', 2)]
        service_raw = row[col_map.get('service', 4)]

        if not project_name_raw or not service_raw:
            continue

        project_name = clean_project_name(project_name_raw)
        project_key = normalize_project_key(project_name)
        service_name = clean_service_name(service_raw)

        if not project_key or not service_name:
            continue

        # Add project if not exists
        if project_key not in projects_dict:
            projects_dict[project_key] = {
                'project_key': project_key,
                'project_name': project_name,
                'project_type': 'yaml',
                'is_public': False,
                'description': '',
            }

        # Build service info
        service_info = {
            'project_key': project_key,
            'service_name': service_name,
            'source': 'template',
            'template_name': row[col_map.get('template_name', 5)] or '',
            'description': row[col_map.get('Description', 6)] or '',
            'namespace': row[col_map.get('Deploy namespace', 14)] or '',
            'auto_sync': True,
            'variable_yaml': '{}',
        }

        services_list.append(service_info)

    projects_list = list(projects_dict.values())
    return projects_list, services_list


def write_standard_excel(projects: list, services: list, output_path: str):
    """Write standard format Excel file with Projects and Services sheets."""

    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create Projects sheet
    ws_projects = wb.create_sheet('Projects')

    # Projects header
    projects_header = ['project_key', 'project_name', 'project_type', 'is_public', 'description']
    ws_projects.append(projects_header)

    # Style header
    for cell in ws_projects[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Projects data
    for project in projects:
        ws_projects.append([
            project['project_key'],
            project['project_name'],
            project['project_type'],
            project['is_public'],
            project['description'],
        ])

    # Auto-fit columns
    for column in ws_projects.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_projects.column_dimensions[column_letter].width = adjusted_width

    # Create Services sheet
    ws_services = wb.create_sheet('Services')

    # Services header - include all variable columns
    services_header = [
        'project_key', 'service_name', 'source', 'template_name',
        'variable_yaml', 'auto_sync', 'description', 'namespace'
    ]

    # Add extra variable columns
    variable_columns = [
        'arch', 'component', 'computeType', 'cpuLimit', 'cpuRequest',
        'envType', 'imagePullPolicy', 'language', 'memoryLimit', 'memoryRequest',
        'metricsEnabled', 'metricsPath', 'nsGatewayHost', 'port', 'probePath',
        'registry', 'tag', 'team'
    ]
    services_header.extend(variable_columns)

    ws_services.append(services_header)

    # Style header
    for cell in ws_services[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Services data
    for service in services:
        row = [
            service['project_key'],
            service['service_name'],
            service['source'],
            service['template_name'],  # Will be filled by user
            service['variable_yaml'],  # JSON string
            service['auto_sync'],
            service['description'],
            service['namespace'],
        ]

        # Add empty cells for variable columns (to be filled by user)
        row.extend([''] * len(variable_columns))

        ws_services.append(row)

    # Auto-fit columns for services sheet
    for column in ws_services.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_services.column_dimensions[column_letter].width = adjusted_width

    # Add Readme sheet with instructions
    ws_readme = wb.create_sheet('Readme')
    readme_content = [
        ['Zadig 批量创建项目标准格式'],
        [''],
        ['说明：'],
        ['1. Projects 表：定义要创建的项目'],
        ['   - project_key: 项目标识（必填，自动生成）'],
        ['   - project_name: 项目名称（必填）'],
        ['   - project_type: 项目类型（固定为 yaml）'],
        ['   - is_public: 是否公开（默认 false）'],
        ['   - description: 项目描述（选填）'],
        [''],
        ['2. Services 表：定义项目下的服务'],
        ['   - project_key: 关联的项目标识（必填）'],
        ['   - service_name: 服务名称（必填）'],
        ['   - source: 创建来源（固定为 template）'],
        ['   - template_name: 服务模板名称（必填，请补充）'],
        ['   - variable_yaml: 服务变量 JSON（选填，或使用右侧列）'],
        ['   - auto_sync: 是否自动同步（默认 true）'],
        ['   - description: 服务描述（选填）'],
        ['   - namespace: 命名空间（预留字段）'],
        [''],
        ['3. 服务变量列（右侧）：'],
        ['   可以直接在这些列中填写值，程序会自动合并到 variable_yaml 中'],
        ['   包括：arch, component, computeType, cpuLimit, cpuRequest,'],
        ['   envType, imagePullPolicy, language, memoryLimit, memoryRequest,'],
        ['   metricsEnabled, metricsPath, nsGatewayHost, port, probePath,'],
        ['   registry, tag, team'],
        [''],
        ['4. 使用方式：'],
        ['   python scripts/zadig_client.py batch-create-projects \\'],
        ['     --input /path/to/standard.xlsx'],
        [''],
        ['注意事项：'],
        ['- template_name 列当前为空，需要手动补充服务模板名称'],
        ['- 如果使用右侧列填写变量，variable_yaml 列可以为空 "{}"'],
        ['- 确保所有服务的 template_name 都已填写后再执行批量创建'],
    ]

    for row in readme_content:
        ws_services.append(row)  # This should be ws_readme

    # Auto-fit readme columns
    ws_readme.column_dimensions['A'].width = 80

    wb.save(output_path)
    print(f"✓ 标准格式文件已生成: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description='Convert OB application list to Zadig standard format'
    )
    parser.add_argument(
        'input',
        help='Input Excel file path (e.g., "Onebackend 资源清单汇总.xlsx")'
    )
    parser.add_argument(
        'output',
        help='Output Excel file path (e.g., "zadig_projects_standard.xlsx")'
    )

    args = parser.parse_args()

    input_path = Path(args.input).expanduser()
    output_path = Path(args.output).expanduser()

    if not input_path.exists():
        print(f"❌ 输入文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    print(f"📖 读取源文件: {input_path}")

    try:
        projects, services = read_source_excel(str(input_path))

        print(f"✓ 提取了 {len(projects)} 个项目")
        print(f"✓ 提取了 {len(services)} 个服务")

        # Show sample projects
        print("\n项目示例：")
        for p in projects[:5]:
            print(f"  - {p['project_key']} ({p['project_name']})")
        if len(projects) > 5:
            print(f"  ... 还有 {len(projects) - 5} 个项目")

        print(f"\n📝 写入标准格式: {output_path}")
        write_standard_excel(projects, services, str(output_path))

        print("\n✅ 转换完成！")
        print(f"\n下一步：")
        print(f"1. 打开 {output_path}")
        print(f"2. 在 Services 表中补充 template_name 列")
        print(f"3. （可选）填写服务变量列")
        print(f"4. 运行批量创建命令")

    except Exception as e:
        print(f"❌ 转换失败: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
